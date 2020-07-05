import glob
import os
from datetime import datetime
from zipfile import ZipFile

import openpyxl
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font, PatternFill

import netgear_aliases as aliases
import netgear_queries as queries
# from dito_logger import DitoLogger
from random import randint

# dl = DitoLogger(os.path.basename(__file__))
# logger = dl.logger


def checkDir(path):
    if not os.path.exists(path):
        os.makedirs(path)


def writeFRI(fri_filepath, major_equipment, manufacturing, equipment_category, equipment_type, unit_category,
             serial_number, web_defined_status, description, mobile_defined_status, remark, site_id_excel,
             barcode_flag, movement_flag, unit_of_measurement, flag_waste, req_qty_db, req_qty):
    if not os.path.exists(fri_filepath):
        text = "text|p1_1_" + serial_number + "|" + major_equipment + \
               "|@|text|p1_2_" + serial_number + "|" + manufacturing + \
               "|@|text|p1_3_" + serial_number + "|" + equipment_category + \
               "|@|text|p1_4_" + serial_number + "|" + equipment_type + \
               "|@|text|p1_5_" + serial_number + "|" + unit_category + \
               "|@|text|p1_6_" + serial_number + "|" + serial_number + \
               "|@|text|p1_7_" + serial_number + "|" + web_defined_status + \
               "|@|text|p1_8_" + serial_number + "|" + description + \
               "|@|text|p1_9_" + serial_number + "|" + mobile_defined_status + \
               "|@|text|p1_10_" + serial_number + "|" + remark + \
               "|@|text|site_id|" + site_id_excel + \
               "|@|text|barcode_flag|" + barcode_flag + \
               "|@|text|movement_flag|" + movement_flag + \
               "|@|text|unit_of_measurement|" + unit_of_measurement + \
               "|@|text|flag_waste|" + flag_waste + \
               "|@|text|req_qty_db|" + req_qty_db + \
               "|@|text|req_qty|" + req_qty + '|'

        f = open(fri_filepath, "w", encoding="utf8")
        f.write(text)
        f.close()

        fri_filename = os.path.basename(fri_filepath)
        print("\tFRI FILE CREATED:\t\t%s" % fri_filename)
    else:
        print("\tFRI FILE ALREADY EXISTS")


def writeKT(kt_filepath, site_id):
    if not os.path.exists(kt_filepath):
        longlat = queries.getLongLat(site_id)[0]
        long, lat = str(longlat[0]), str(longlat[1])

        text = "gps_latlong|" + long + "," + lat + "|@|gps_time|2018-07-27 15:25:26|@|networkoperator|51011|@|mcc|510|@|mnc|11|@|celllocation|[22027,56433158,0]|@|cid|56433158|@|lac|22027|@|networktype|LTE|@|cid_1|-1|@|lac_1|-1|@|rssi_1|9 dBm|@|cid_2|-1|@|lac_2|-1|@|rssi_2|29 dBm|@|cid_3|-1|@|lac_3|-1|@|rssi_3|29 dBm|@|cid_4|-1|@|lac_4|-1|@|rssi_4|31 dBm"
        f = open(kt_filepath, "w")
        f.write(text)
        f.close()

        kt_filename = os.path.basename(kt_filepath)
        print("\tKT FILE CREATED:\t\t%s" % kt_filename)
    else:
        print("\tKT FILE ALREADY EXISTS")


def writeWPD(wpd_filepath, kt_filepath):
    if not os.path.exists(wpd_filepath):
        zipObj = ZipFile(wpd_filepath, 'w')

        zipObj.write(kt_filepath, os.path.basename(kt_filepath))

        zipObj.close()

        wpd_filename = os.path.basename(wpd_filepath)
        print("\tWPD FILE CREATED:\t\t%s" % wpd_filename)
    else:
        print("\tWPD FILE ALREADY EXISTS")


def processATF(excelfile, output_path, now_date):
    print('PROCESSING DISPOSAL :\t', os.path.basename(excelfile))

    book = openpyxl.load_workbook(excelfile)
    sheet = book['List to Inbound to DHL Cikarang']

    start_row = 1
    end_row = sheet.max_row

    material_list = []
    duplicate_serialnumbers = []
    duplicate_serialnumbers_counter = 1

    rejected_tasks = []

    for row in range(start_row + 1, end_row + 1):
        task_id = queries.generateTaskIDDisposal()[0][0]
        serial_number_xls = str(sheet['H' + str(row)].value).strip()

        print("\tPROCESSING (%s/%s): SN %s of task %s" % (row - 1, end_row + 1, serial_number_xls, task_id))

        ne_code = str(sheet['C' + str(row)].value).strip()

        major_equipment_xls = '-'
        unit_category_xls = str(sheet['D' + str(row)].value).strip()

        netgearmaterial = queries.getPN(ne_code, major_equipment_xls, unit_category_xls)

        major_equipment = netgearmaterial[0][2]
        manufacturing = netgearmaterial[0][3]
        equipment_category = netgearmaterial[0][4]
        equipment_type = netgearmaterial[0][5]
        unit_category = unit_category_xls
        flag_waste = netgearmaterial[0][7]
        unit_of_measurement = netgearmaterial[0][8]

        task_id_folder = output_path + "\\" + task_id
        checkDir(task_id_folder)

        site_id = str(sheet['B' + str(row)].value).strip()

        web_defined_status = ' '
        description = ' '
        mobile_defined_status = 'NOT EXIST'
        remark = ' '
        barcode_flag = ' '
        movement_flag = 'MOVING'
        req_qty_db = str(sheet['K' + str(row)].value).strip()
        req_qty = str(sheet['K' + str(row)].value).strip()
        # reuse_matrix = ' '
        # pallet_id = ' ',
        # location_id = ' ',
        # condition = ' ',
        # remark_condition = ' '

        exception_serialnumbers_in = ['broken', 'blank', 'none', 'terbaca', 'buram', 'Broken']
        exception_serialnumbers_is = ['NA', '-', 'NO SN']

        if serial_number_xls is None:
            serial_number = queries.getDummySN()[0][0]
        elif any(sn in serial_number_xls.lower() for sn in exception_serialnumbers_in):
            serial_number = queries.getDummySN()[0][0]
        elif any(sn == serial_number_xls.upper() for sn in exception_serialnumbers_is):
            serial_number = queries.getDummySN()[0][0]
        elif any(sn == serial_number_xls.lower() for sn in duplicate_serialnumbers):
            serial_number = serial_number_xls.lower() + ' (%s)' % str(duplicate_serialnumbers_counter)
            duplicate_serialnumbers_counter += 1
        else:
            serial_number = serial_number_xls

        # =============== FILES FOR TASK ORIGIN =================

        fri_filepath = task_id_folder + "\\" + "%s_%s_%s.fri" % (task_id, site_id, serial_number)
        fri_content = (major_equipment, manufacturing, equipment_category, equipment_type, unit_category_xls,
                       serial_number, web_defined_status, description, mobile_defined_status, remark,
                       site_id, barcode_flag, movement_flag, unit_of_measurement, flag_waste,
                       req_qty_db, req_qty)

        writeFRI(fri_filepath, *fri_content)

        # kt_filepath = task_id_folder + "\\" + "%s_%s_%s.kt" % (task_id, site_id_fri, serial_number)
        # writeKT(kt_filepath, site_id_fri)
        #
        # wpd_filepath = task_id_folder + "\\" + "%s_%s_%s.wpd" % (task_id, site_id_fri, serial_number)
        # writeWPD(wpd_filepath, kt_filepath)

        item_id_date = now_date.replace('_', '')
        item_id = queries.generateItemID(item_id_date)[0][0]

        tbl_item = (
            major_equipment, manufacturing, equipment_category, equipment_type, unit_category,
            serial_number, unit_of_measurement, req_qty, flag_waste, now_date, movement_flag,
            mobile_defined_status, task_id, item_id
        )
        queries.insertTblItem(*tbl_item)
        queries.insertTblItemFRI(*tbl_item)

        duplicate_serialnumbers.append(serial_number_xls.lower())

        # RECORD MATERIAL

        system = '-'
        manufacturing_xls = str(sheet['E' + str(row)].value).strip()
        equipment_category_xls = str(sheet['F' + str(row)].value).strip()
        equipment_type_xls = str(sheet['G' + str(row)].value).strip()
        flag_waste_xls = '-'
        unit_of_measurement_xls = str(sheet['L' + str(row)].value).strip()

        data_xls_to_record = (
            system, major_equipment, manufacturing_xls, equipment_category_xls, equipment_type_xls, unit_category,
            serial_number_xls, flag_waste_xls, unit_of_measurement_xls, req_qty
        )

        kd_acs = netgearmaterial[0][0] if netgearmaterial else ''
        site_id_destination = '03WDHL02'

        data_netgear_to_record = (
            kd_acs, ne_code, major_equipment, manufacturing, equipment_category, equipment_type, unit_category,
            serial_number, flag_waste, unit_of_measurement, req_qty, now_date, movement_flag, mobile_defined_status,
            task_id, os.path.basename(excelfile), site_id, site_id_destination
        )

        if (site_id != str(sheet['B' + str(row + 1)].value).strip() or ne_code != str(sheet['C' + str(row + 1)].value).strip()):
            project_id = "ZTE Movement (Ex-MPFS)"
            plan_date = '2099-12-31'
            status_origin = 'Need Approval2'
            task_type = 'Movement+Collection'
            remark = "Disposal ZTE Movement (Ex-MPFS) Backend Upload"
            division = aliases.getDivisionAlias(queries.getSite(site_id)[0][1])
            last_update = now_date
            vendor = 'PT.ZTE Indonesia'
            atf_no = '-'

            ne_id = queries.getNEID(site_id, ne_code)[0][0]

            task_origin_data = (
                task_id, project_id, site_id, plan_date, status_origin,
                remark, task_type, last_update, vendor, ne_id,
                atf_no, division, site_id_destination
            )
            queries.insertTblTask(*task_origin_data)

        material_list.append((data_xls_to_record, data_netgear_to_record))

    return material_list, rejected_tasks


def recordMaterial(path, atfmaterials, rejectedatfs):
    wb = Workbook()

    ws = wb.active
    ws.title = 'Input vs Output'

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ws['A1'] = 'INPUT'
    ws.merge_cells('A1:J1')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].font = Font(size=18, bold=True)
    ws['A1'].border = thin_border

    atf_cell_color = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
    ws['A2'] = 'System'
    ws['A2'].fill = atf_cell_color
    ws['A2'].border = thin_border
    ws['B2'] = 'Subsystem'
    ws['B2'].fill = atf_cell_color
    ws['B2'].border = thin_border
    ws['C2'] = 'Brand'
    ws['C2'].fill = atf_cell_color
    ws['C2'].border = thin_border
    ws['D2'] = 'Equipment Type'
    ws['D2'].fill = atf_cell_color
    ws['D2'].border = thin_border
    ws['E2'] = 'Description'
    ws['E2'].fill = atf_cell_color
    ws['E2'].border = thin_border
    ws['F2'] = 'Product Code'
    ws['F2'].fill = atf_cell_color
    ws['F2'].border = thin_border
    ws['G2'] = 'Serial Number'
    ws['G2'].fill = atf_cell_color
    ws['G2'].border = thin_border
    ws['H2'] = 'Flag Waste'
    ws['H2'].fill = PatternFill(start_color='ADD8E6', end_color='87CEEB', fill_type='solid')
    ws['H2'].border = thin_border
    ws['I2'] = 'UoM'
    ws['I2'].fill = atf_cell_color
    ws['I2'].border = thin_border
    ws['J2'] = 'Qty'
    ws['J2'].fill = atf_cell_color
    ws['J2'].border = thin_border

    ws['K1'] = 'OUTPUT'
    ws.merge_cells('K1:Y1')
    ws['K1'].alignment = Alignment(horizontal='center')
    ws['K1'].font = Font(size=18, bold=True)
    ws['K1'].border = thin_border

    netgear_cell_color = PatternFill(start_color='FED8B1', end_color='90EE90', fill_type='solid')

    ws['K2'] = 'Kode Predefine'
    ws['K2'].fill = netgear_cell_color
    ws['K2'].border = thin_border
    ws['L2'] = 'NE Code'
    ws['L2'].fill = netgear_cell_color
    ws['L2'].border = thin_border
    ws['M2'] = 'Major Equipment'
    ws['M2'].fill = netgear_cell_color
    ws['M2'].border = thin_border
    ws['N2'] = 'Manufacturing'
    ws['N2'].fill = netgear_cell_color
    ws['N2'].border = thin_border
    ws['O2'] = 'Equipment Category'
    ws['O2'].fill = netgear_cell_color
    ws['O2'].border = thin_border
    ws['P2'] = 'Equipment Type'
    ws['P2'].fill = netgear_cell_color
    ws['P2'].border = thin_border
    ws['Q2'] = 'Part Number'
    ws['Q2'].fill = netgear_cell_color
    ws['Q2'].border = thin_border
    ws['R2'] = 'Serial Number'
    ws['R2'].fill = netgear_cell_color
    ws['R2'].border = thin_border
    ws['S2'] = 'Flag Waste'
    ws['S2'].fill = netgear_cell_color
    ws['S2'].border = thin_border
    ws['T2'] = 'UoM'
    ws['T2'].fill = netgear_cell_color
    ws['T2'].border = thin_border
    ws['U2'] = 'QTY'
    ws['U2'].fill = netgear_cell_color
    ws['U2'].border = thin_border
    ws['V2'] = 'Last Update'
    ws['V2'].fill = netgear_cell_color
    ws['V2'].border = thin_border
    ws['W2'] = 'Movement Status'
    ws['W2'].fill = netgear_cell_color
    ws['W2'].border = thin_border
    ws['X2'] = 'Status'
    ws['X2'].fill = netgear_cell_color
    ws['X2'].border = thin_border
    ws['Y2'] = 'Task ID'
    ws['Y2'].fill = netgear_cell_color
    ws['Y2'].border = thin_border

    ws['Z1'] = 'Site ID Origin'
    ws.merge_cells('Z1:Z2')
    ws['Z1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['Z1'].border = thin_border
    ws['Z1'].font = Font(size=11, bold=True)

    ws['AA1'] = 'Site ID Destination'
    ws.merge_cells('AA1:AA2')
    ws['AA1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['AA1'].border = thin_border
    ws['AA1'].font = Font(size=11, bold=True)

    counter_atf = 3
    for atfmaterial in atfmaterials:
        ws['A' + str(counter_atf)] = atfmaterial[0][0]  # ATF - SYSTEM
        ws['A' + str(counter_atf)].border = thin_border
        ws['B' + str(counter_atf)] = atfmaterial[0][1]  # ATF - SUBSYSTEM
        ws['B' + str(counter_atf)].border = thin_border
        ws['C' + str(counter_atf)] = atfmaterial[0][2]  # ATF - BRAND
        ws['C' + str(counter_atf)].border = thin_border
        ws['D' + str(counter_atf)] = atfmaterial[0][3]  # ATF - EQUIPMENT TYPE
        ws['D' + str(counter_atf)].border = thin_border
        ws['E' + str(counter_atf)] = atfmaterial[0][4]  # ATF - DESCRIPTION
        ws['E' + str(counter_atf)].border = thin_border
        ws['F' + str(counter_atf)] = atfmaterial[0][5]  # ATF - PRODUCT CODE
        ws['F' + str(counter_atf)].border = thin_border
        ws['G' + str(counter_atf)] = atfmaterial[0][6]  # ATF - SERIAL NUMBER
        ws['G' + str(counter_atf)].border = thin_border
        ws['H' + str(counter_atf)] = atfmaterial[0][7]  # ATF - FLAG WASTE
        ws['H' + str(counter_atf)].border = thin_border
        ws['I' + str(counter_atf)] = atfmaterial[0][8]  # ATF - UOM
        ws['I' + str(counter_atf)].border = thin_border
        ws['J' + str(counter_atf)] = atfmaterial[0][9]  # ATF - QTY
        ws['J' + str(counter_atf)].border = thin_border

        grey = PatternFill(start_color='CDCDCD', end_color='FFFF00', fill_type='solid')
        yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws['K' + str(counter_atf)] = atfmaterial[1][0]  # NETGEAR - KODE PREDEFINE
        ws['K' + str(counter_atf)].border = thin_border
        if not atfmaterial[1][0]:
            ws['K' + str(counter_atf)].fill = yellow
        else:
            ws['K' + str(counter_atf)].fill = grey

        ws['L' + str(counter_atf)] = atfmaterial[1][1]  # NETGEAR - NE CODE
        ws['L' + str(counter_atf)].border = thin_border
        ws['M' + str(counter_atf)] = atfmaterial[1][2]  # NETGEAR - MAJOR EQUIPMENT
        ws['M' + str(counter_atf)].border = thin_border
        ws['N' + str(counter_atf)] = atfmaterial[1][3]  # NETGEAR - MANUFACTURER
        ws['N' + str(counter_atf)].border = thin_border
        ws['O' + str(counter_atf)] = atfmaterial[1][4]  # NETGEAR - EQUIPMENT CATEGORY
        ws['O' + str(counter_atf)].border = thin_border
        ws['P' + str(counter_atf)] = atfmaterial[1][5]  # NETGEAR - EQUIPMENT TYPE
        ws['P' + str(counter_atf)].border = thin_border
        ws['Q' + str(counter_atf)] = atfmaterial[1][6]  # NETGEAR - PART NUMBER
        ws['Q' + str(counter_atf)].border = thin_border
        ws['R' + str(counter_atf)] = atfmaterial[1][7]  # NETGEAR - SERIAL NUMBER
        ws['R' + str(counter_atf)].border = thin_border
        ws['S' + str(counter_atf)] = atfmaterial[1][8]  # NETGEAR - FLAG WASTE
        ws['S' + str(counter_atf)].border = thin_border
        ws['T' + str(counter_atf)] = atfmaterial[1][9]  # NETGEAR - UOM
        ws['T' + str(counter_atf)].border = thin_border
        ws['U' + str(counter_atf)] = atfmaterial[1][10]  # NETGEAR - QTY
        ws['U' + str(counter_atf)].border = thin_border
        ws['V' + str(counter_atf)] = atfmaterial[1][11]  # NETGEAR - LAST UPDATE
        ws['V' + str(counter_atf)].border = thin_border
        ws['W' + str(counter_atf)] = atfmaterial[1][12]  # NETGEAR - MOVEMENT STATUS
        ws['W' + str(counter_atf)].border = thin_border
        ws['X' + str(counter_atf)] = atfmaterial[1][13]  # NETGEAR - STATUS
        ws['X' + str(counter_atf)].border = thin_border
        ws['Y' + str(counter_atf)] = atfmaterial[1][14]  # NETGEAR - TASK ID
        ws['Y' + str(counter_atf)].border = thin_border

        light_green = PatternFill(start_color='ABFE01', end_color='ABFE01', fill_type='solid')
        ws['Z' + str(counter_atf)] = atfmaterial[1][16]
        ws['Z' + str(counter_atf)].border = thin_border
        ws['Z' + str(counter_atf)].fill = light_green
        ws['Z' + str(counter_atf)].font = Font(size=11, bold=True)

        ws['AA' + str(counter_atf)] = atfmaterial[1][17]
        ws['AA' + str(counter_atf)].border = thin_border
        ws['AA' + str(counter_atf)].fill = light_green
        ws['AA' + str(counter_atf)].font = Font(size=11, bold=True)

        red = PatternFill(start_color='FF6961', end_color='FF6961', fill_type='solid')
        green = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        if atfmaterial[0][2].lower() != atfmaterial[1][3].lower():  # BRAND VS MANUFACTURER
            ws['C' + str(counter_atf)].fill = red
            ws['N' + str(counter_atf)].fill = green
        if atfmaterial[0][3].lower() != atfmaterial[1][4].lower():  # EQ. TYPE VS EQ. CATEGORY
            ws['D' + str(counter_atf)].fill = red
            ws['O' + str(counter_atf)].fill = green
        if atfmaterial[0][4].lower() != atfmaterial[1][5].lower():  # DESCRIPTION VS EQ. TYPE
            ws['E' + str(counter_atf)].fill = red
            ws['P' + str(counter_atf)].fill = green
        if atfmaterial[0][2].lower() != atfmaterial[1][3].lower():  # SN ATF VS SN NETGEAR
            ws['G' + str(counter_atf)].fill = red
            ws['R' + str(counter_atf)].fill = green
        if atfmaterial[0][7].lower() != atfmaterial[1][8].lower():  # FLAG WASTE ATF VS FLAG WASTE NETGEAR
            ws['H' + str(counter_atf)].fill = red
            ws['S' + str(counter_atf)].fill = green
        if atfmaterial[0][8].lower() != atfmaterial[1][9].lower():  # UOM ATF VS UOM NETGEAR
            ws['I' + str(counter_atf)].fill = red
            ws['T' + str(counter_atf)].fill = green

        counter_atf += 1

    ws2 = wb.create_sheet('Rejected Input')
    ws2['A1'] = 'ATF Filename'
    ws2['A1'].border = thin_border
    ws2['A1'].font = Font(size=18, bold=True)
    ws2['B1'] = 'Reason'
    ws2['B1'].border = thin_border
    ws2['B1'].font = Font(size=18, bold=True)

    counter_reject = 2
    for rejectedatf in rejectedatfs:
        red = PatternFill(start_color='FF6961', end_color='FF6961', fill_type='solid')

        ws2['A' + str(counter_reject)] = rejectedatf[0]  # NE CODE
        ws2['A' + str(counter_reject)].border = thin_border
        ws2['A' + str(counter_reject)].font = Font(size=11, bold=True)
        ws2['B' + str(counter_reject)] = rejectedatf[1]  # NE CODE
        ws2['B' + str(counter_reject)].border = thin_border
        ws2['B' + str(counter_reject)].fill = red

        counter_reject += 1

    wb.save(path)


if __name__ == '__main__':
    nowdatetime = datetime.today().strftime('%Y%m%d_%H%M%S')
    batch_name = 'BATCH ' + nowdatetime
    print('RUNNING ' + batch_name)

    input_path = "D:\\KERJAAN\\SCRIPTS\\SCRIPTS\\Made by AMG\\MANUAL - NETgear ATF Uploader\\INPUT\\DISPOSALDHL"
    list_of_atf_excess = glob.glob("%s\\*.xlsx" % input_path)

    output_path = "D:\\KERJAAN\\SCRIPTS\\SCRIPTS\\Made by AMG\\MANUAL - NETgear ATF Uploader\\OUTPUT\\DISPOSALDHL"
    output_path = output_path + '\\' + batch_name
    checkDir(output_path)

    done_path = "D:\\KERJAAN\\SCRIPTS\\SCRIPTS\\Made by AMG\\MANUAL - NETgear ATF Uploader\\DONE\\DISPOSALDHL"
    done_path = done_path + '\\' + batch_name
    checkDir(done_path)

    material_path = "D:\\KERJAAN\\SCRIPTS\\SCRIPTS\\Made by AMG\\MANUAL - NETgear ATF Uploader\\MATERIALS\\DISPOSALDHL"
    material_path = material_path + '\\' + batch_name + '.xlsx'

    material_list = []
    rejected_tasks = []

    for atf_file in list_of_atf_excess:
        atf_materials = processATF(atf_file, output_path, nowdatetime)
        if atf_materials:
            material_list = material_list + atf_materials[0]
            rejected_tasks = rejected_tasks + atf_materials[1]

            atf_done_path = done_path + '\\' + os.path.basename(atf_file)
            os.rename(atf_file, atf_done_path)

    recordMaterial(material_path, material_list, rejected_tasks)
